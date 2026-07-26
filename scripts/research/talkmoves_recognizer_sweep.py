#!/usr/bin/env python3
"""Run the strategy recognizer over every student turn in the TalkMoves corpus.

The recognizer models 106 execution-observed automata and abstains by design.
Until now it has met three transcripts and one worked demonstration.  The Sumner
lab's released corpus is 567 K-12 mathematics lesson transcripts, sentence
segmented, with teacher and student moves already coded by people — real
classroom talk at a scale this instrument has never been asked to face.

What this measures is deliberately narrow, because it is what can be measured
without a second labelling effort.  TalkMoves labels *discourse* moves; Hermes
recognizes *mathematical strategies*.  Those are different taxonomies and no
agreement statistic between them would mean anything.  What the corpus can settle
is the question a prototype has to answer before anyone funds a study:

  When this instrument reads real classroom speech, how often does it say
  something, how often does it stay silent, and is what it says defensible?

Silence is the interesting half.  A recognizer that fires on everything is a
keyword matcher wearing a citation; one that never fires is furniture.  The rate
and its distribution over machines are reported, never a claim that a recognition
is correct — correctness needs a reader and that is a separate, later pass over
the sample this writes out.

**Licence.** TalkMoves is CC BY-NC-SA 4.0 (Jacobs et al. 2022; Suresh et al.
2021).  Under the standing rights ruling this driver writes **derived counts**
and no transcript text.  A `--sample` file of short fragments can be written for
a human reading pass, and it is kept out of the repository tree by default.

``--mode episode`` groups all student utterances from each transcript in corpus
order.  It measures whether carrying an automaton frontier across sentence
boundaries assembles longer ordered traces than the stored sentence baseline.
"""
from __future__ import annotations

import argparse
import collections
import json
import re
import subprocess
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[2]
CORPUS = Path.home() / "Documents/GitHub/TalkMoves/data"
SPLITS = ("train_data_504.xlsx", "test_data_63.xlsx")
# pandas.read_excel treats these strings as missing by default.  The fallback
# must do the same or it admits 902 literal "nan" rows and six literal "None"
# rows that are absent from the established 58,926-sentence corpus population.
PANDAS_DEFAULT_NA_VALUES = frozenset({
    "", "#N/A", "#N/A N/A", "#NA", "-1.#IND", "-1.#QNAN", "-NaN", "-nan",
    "1.#IND", "1.#QNAN", "<NA>", "N/A", "NA", "NULL", "NaN", "None",
    "n/a", "nan", "null",
})

# Short mathematical constructions to count only when the recognizer abstains.
# These are analysis probes, not recognition surfaces.  The groups keep
# inflection and pronoun variants together so the output reports a divergence
# class rather than copying transcript sentences into the tree.
DIVERGENCE_PATTERNS: tuple[
    tuple[str, str, tuple[tuple[str, ...], ...]], ...
] = (
    ("deictic_addition", "combine_quantities",
     (("add", "it"), ("added", "it"), ("adding", "it"))),
    ("take_away_process", "remove_quantity",
     (("take", "away"), ("took", "away"), ("taking", "away"))),
    ("division_operator", "compute_quotient",
     (("divide", "by"), ("divided", "by"), ("dividing", "by"))),
    ("deictic_division", "compute_quotient",
     (("divide", "it", "by"), ("divided", "it", "by"),
      ("dividing", "it", "by"))),
    ("multiplication_operator", "compute_product",
     (("multiply", "by"), ("multiplied", "by"), ("multiplying", "by"))),
    ("count_by_process", "iterate_composite_unit",
     (("count", "by"), ("counted", "by"), ("counting", "by"))),
    ("operand_decomposition", "decompose_operand",
     (("split", "it", "up"), ("split", "them", "up"),
      ("break", "it", "up"), ("broke", "it", "up"),
      ("break", "apart"), ("broke", "apart"))),
    ("common_denominator", "align_to_common_unit",
     (("common", "denominator"), ("same", "denominator"))),
    ("equal_groups", "replicate_equal_groups",
     (("equal", "groups"),)),
    ("equal_pieces", "partition_into_equal_parts",
     (("equal", "pieces"), ("same", "size", "pieces"))),
    ("number_line_reference", "establish_reference_frame",
     (("number", "line"),)),
    ("magnitude_relation", "compare_magnitudes",
     (("greater", "than"), ("less", "than"))),
    ("totalizing_phrase", "accumulate_total",
     (("all", "together"), ("in", "all"))),
    ("measurement_division", "measure_out_group_size",
     (("goes", "into"), ("fit", "into"), ("fits", "into"))),
)

# One SWI process reads sentences from stdin and prints recognitions, because
# a process per sentence would put a fork between the instrument and every
# clause of classroom speech.
PROLOG_DRIVER = r"""
:- use_module(hermes(strategy_recognizer), [recognize_strategies/2]).

% Run with -g main -t halt, never initialization(main, main): under -l/-s the
% toplevel otherwise reads the piped sentences as Prolog source and dies on the
% first clause of English. The same trap silently disabled two checks earlier
% today.
main :-
    read_line_to_string(user_input, First),
    loop(First).

loop(end_of_file) :- !.
loop(Line) :-
    (   catch(recognize_strategies(Line, Candidates), _, fail)
    ->  true
    ;   Candidates = []
    ),
    emit(Candidates),
    read_line_to_string(user_input, Next),
    loop(Next).

emit([]) :- !, format("-~n").
emit(Candidates) :-
    findall(Text,
            ( member(C, Candidates), candidate_text(C, Text) ),
            Texts),
    atomic_list_concat(Texts, ' ', Joined),
    format("~w~n", [Joined]).

% A candidate is a dict. What matters for a sweep is which machine, how much of
% it was matched, and on what basis: support_level lexical_hint with
% matched_count 1 means one token met one action, which is a hint and not a
% reading, and reporting it beside a partial_trace would inflate the result.
candidate_text(Candidate, Text) :-
    get_dict(operation, Candidate, Operation),
    get_dict(kind, Candidate, Kind),
    ( get_dict(support_level, Candidate, Support) -> true ; Support = unknown ),
    ( get_dict(matched_count, Candidate, Matched) -> true ; Matched = 0 ),
    ( get_dict(expected_actions, Candidate, Expected) -> true ; Expected = 0 ),
    ( get_dict(confidence, Candidate, Confidence) -> true ; Confidence = 0 ),
    format(atom(Text), '~w/~w|~w|~w|~w|~4f',
           [Operation, Kind, Support, Matched, Expected, Confidence]).
"""

PROLOG_EPISODE_DRIVER = r"""
:- use_module(library(http/json), [atom_json_term/3]).
:- use_module(hermes(strategy_recognizer),
              [recognize_strategy_episode/2]).

main :-
    read_line_to_string(user_input, First),
    loop(First).

loop(end_of_file) :- !.
loop(Line) :-
    (   catch(
            ( atom_json_term(Line, Utterances,
                             [value_string_as(string)]),
              recognize_strategy_episode(Utterances, Candidates)
            ),
            _,
            fail)
    ->  true
    ;   Candidates = []
    ),
    emit(Candidates),
    read_line_to_string(user_input, Next),
    loop(Next).

emit([]) :- !, format("-~n").
emit(Candidates) :-
    findall(Text,
            ( member(C, Candidates), candidate_text(C, Text) ),
            Texts),
    atomic_list_concat(Texts, ' ', Joined),
    format("~w~n", [Joined]).

candidate_text(Candidate, Text) :-
    get_dict(operation, Candidate, Operation),
    get_dict(kind, Candidate, Kind),
    get_dict(support_level, Candidate, Support),
    get_dict(matched_count, Candidate, Matched),
    get_dict(expected_actions, Candidate, Expected),
    get_dict(confidence, Candidate, Confidence),
    get_dict(ordered_action_count, Candidate, Ordered),
    get_dict(current_frontier, Candidate, Frontier),
    get_dict(status, Frontier, Status),
    format(atom(Text), '~w/~w|~w|~w|~w|~4f|~w|~w',
           [Operation, Kind, Support, Matched, Expected, Confidence,
            Ordered, Status]).
"""


def student_sentences(limit_transcripts: int) -> list[tuple[str, str, str]]:
    """(transcript, speaker, sentence) for student turns, from both splits."""
    try:
        import pandas
    except ImportError:
        pandas = None
    rows: list[tuple[str, str, str]] = []
    seen: set[str] = set()
    for split in SPLITS:
        path = CORPUS / split
        if not path.exists():
            print(f"FAIL: {path} not found", file=sys.stderr)
            raise SystemExit(1)
        if pandas is not None:
            frame = pandas.read_excel(path)
            workbook_rows = zip(
                frame["Transcript"], frame["Speaker"], frame["Sentence"])
        else:
            try:
                from openpyxl import load_workbook
            except ImportError:
                print("FAIL: pandas or openpyxl is needed to read the "
                      "TalkMoves workbooks", file=sys.stderr)
                raise SystemExit(1)
            workbook = load_workbook(path, read_only=True, data_only=True)
            sheet = workbook.active
            iterator = sheet.iter_rows(values_only=True)
            header = next(iterator)
            columns = {
                str(value): index
                for index, value in enumerate(header)
                if value is not None
            }
            required = {"Transcript", "Speaker", "Sentence"}
            missing = sorted(required - columns.keys())
            if missing:
                workbook.close()
                print(f"FAIL: {path} lacks column(s): {', '.join(missing)}",
                      file=sys.stderr)
                raise SystemExit(1)
            workbook_rows = (
                (values[columns["Transcript"]],
                 values[columns["Speaker"]],
                 values[columns["Sentence"]])
                for values in iterator
            )
        for transcript, speaker, sentence in workbook_rows:
            if (not isinstance(sentence, str)
                    or sentence.strip() in PANDAS_DEFAULT_NA_VALUES):
                continue
            name = str(transcript)
            if limit_transcripts and name not in seen and len(seen) >= limit_transcripts:
                continue
            seen.add(name)
            if str(speaker).strip().upper().startswith("S"):
                rows.append((name, str(speaker), sentence.strip()))
        if pandas is None:
            workbook.close()
    return rows


def normalized_words(sentence: str) -> tuple[str, ...]:
    """Lower-case words and numbers for derived pattern counts."""
    return tuple(re.findall(r"[a-z]+|\d+(?:\.\d+)?", sentence.lower()))


def contains_words(tokens: tuple[str, ...], phrase: tuple[str, ...]) -> bool:
    """True when Phrase is a contiguous span of Tokens."""
    width = len(phrase)
    return any(
        tokens[start:start + width] == phrase
        for start in range(len(tokens) - width + 1)
    )


def divergence_matches(sentence: str) -> list[tuple[str, str]]:
    """Named mathematical-language groups present in an abstained sentence."""
    tokens = normalized_words(sentence)
    return [
        (name, action)
        for name, action, variants in DIVERGENCE_PATTERNS
        if any(contains_words(tokens, variant) for variant in variants)
    ]


def recognize(sentences: list[str], timeout: int) -> list[str]:
    driver = ROOT / "scripts/research/_talkmoves_sweep_driver.pl"
    driver.write_text(PROLOG_DRIVER, encoding="utf-8")
    try:
        payload = "\n".join(s.replace("\n", " ") for s in sentences) + "\n"
        result = subprocess.run(
            ["swipl", "-q", "--on-warning=status", "--on-error=status",
             "-l", "paths.pl", "-s", str(driver), "-g", "main", "-t", "halt"],
            cwd=ROOT, input=payload, capture_output=True, text=True,
            timeout=timeout)
    finally:
        driver.unlink(missing_ok=True)
    if result.returncode:
        print(f"FAIL: recognizer exited {result.returncode}\n{result.stderr[-1500:]}",
              file=sys.stderr)
        raise SystemExit(1)
    lines = result.stdout.splitlines()
    if len(lines) != len(sentences):
        print(f"FAIL: {len(lines)} results for {len(sentences)} sentences",
              file=sys.stderr)
        raise SystemExit(1)
    return lines


def recognize_episodes(episodes: list[list[str]], timeout: int) -> list[str]:
    """Recognize one ordered student-utterance sequence per transcript."""
    driver = ROOT / "scripts/research/_talkmoves_episode_sweep_driver.pl"
    driver.write_text(PROLOG_EPISODE_DRIVER, encoding="utf-8")
    try:
        payload = "\n".join(json.dumps(episode) for episode in episodes) + "\n"
        result = subprocess.run(
            ["swipl", "-q", "--on-warning=status", "--on-error=status",
             "-l", "paths.pl", "-s", str(driver), "-g", "main", "-t", "halt"],
            cwd=ROOT, input=payload, capture_output=True, text=True,
            timeout=timeout)
    finally:
        driver.unlink(missing_ok=True)
    if result.returncode:
        print(f"FAIL: episode recognizer exited {result.returncode}\n"
              f"{result.stderr[-1500:]}", file=sys.stderr)
        raise SystemExit(1)
    lines = result.stdout.splitlines()
    if len(lines) != len(episodes):
        print(f"FAIL: {len(lines)} results for {len(episodes)} episodes",
              file=sys.stderr)
        raise SystemExit(1)
    return lines


def episode_payload(
        rows: list[tuple[str, str, str]],
        baseline: dict,
        batch: int,
        timeout: int) -> dict:
    """Derived episode counts, with the stored sentence baseline alongside."""
    grouped: dict[str, list[str]] = {}
    for transcript, _speaker, sentence in rows:
        grouped.setdefault(transcript, []).append(sentence)

    by_support: collections.Counter = collections.Counter()
    by_machine: collections.Counter = collections.Counter()
    three_or_more: collections.Counter = collections.Counter()
    accepting: collections.Counter = collections.Counter()
    episodes_with_candidates = 0
    candidate_instances = 0
    items = list(grouped.items())

    for start in range(0, len(items), batch):
        chunk = items[start:start + batch]
        results = recognize_episodes(
            [utterances for _, utterances in chunk], timeout)
        for (_transcript, _utterances), line in zip(chunk, results):
            if line.strip() == "-":
                continue
            episodes_with_candidates += 1
            for token in line.split():
                parts = token.split("|")
                if len(parts) != 7 or "/" not in parts[0]:
                    continue
                (machine, support, _matched, _expected, _confidence,
                 ordered, status) = parts
                ordered_count = int(ordered)
                candidate_instances += 1
                by_machine[machine] += 1
                by_support[support] += 1
                if ordered_count >= 3:
                    three_or_more[machine] += 1
                if status == "accepting":
                    accepting[machine] += 1
        print(f"  {min(start + batch, len(items))}/{len(items)}"
              f"  episodes={episodes_with_candidates}"
              f"  ordered>=3={sum(three_or_more.values())}"
              f"  accepting={sum(accepting.values())}",
              flush=True)

    lexical_hints = by_support["lexical_hint"]
    partial_traces = by_support["partial_trace"]
    episode_ratio = (
        round(partial_traces / lexical_hints, 6)
        if lexical_hints else None
    )
    sentence_support = baseline.get("candidate_support_levels", {})
    sentence_ratio = baseline.get("partial_trace_to_lexical_hint_ratio")
    if sentence_ratio is None and sentence_support.get("lexical_hint"):
        sentence_ratio = round(
            sentence_support.get("partial_trace", 0)
            / sentence_support["lexical_hint"],
            6)

    return {
        "recognition_unit":
            "all student utterances in one transcript, in corpus order",
        "episode_count": len(grouped),
        "student_utterances": len(rows),
        "episodes_with_a_candidate": episodes_with_candidates,
        "episode_abstention_rate":
            round(1 - episodes_with_candidates / len(grouped), 4),
        "candidate_instances": candidate_instances,
        "distinct_machines_fired": len(by_machine),
        "candidate_support_levels": dict(by_support.most_common()),
        "partial_trace_to_lexical_hint_ratio": episode_ratio,
        "machine_episode_candidates_reaching_three_or_more_ordered_actions":
            sum(three_or_more.values()),
        "distinct_machines_reaching_three_or_more_ordered_actions":
            len(three_or_more),
        "top_machines_reaching_three_or_more_ordered_actions":
            dict(three_or_more.most_common(25)),
        "machine_episode_candidates_reaching_accepting_state":
            sum(accepting.values()),
        "distinct_machines_reaching_accepting_state": len(accepting),
        "top_machines_reaching_accepting_state":
            dict(accepting.most_common(25)),
        "comparison_to_sentence_baseline": {
            "sentence_population": {
                "student_sentences": baseline.get("student_sentences"),
                "transcripts": baseline.get("transcripts"),
            },
            "sentence_candidate_support_levels": {
                "lexical_hint": sentence_support.get("lexical_hint", 0),
                "partial_trace": sentence_support.get("partial_trace", 0),
            },
            "episode_candidate_support_levels": {
                "lexical_hint": lexical_hints,
                "partial_trace": partial_traces,
            },
            "sentence_partial_trace_to_lexical_hint_ratio": sentence_ratio,
            "episode_partial_trace_to_lexical_hint_ratio": episode_ratio,
        },
    }


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--mode", choices=("sentence", "episode"),
                        default="sentence")
    parser.add_argument("--transcripts", type=int, default=0,
                        help="0 = every transcript in both splits")
    parser.add_argument("--batch", type=int, default=2000)
    parser.add_argument("--timeout", type=int, default=1800)
    parser.add_argument("--output", type=Path,
                        default=ROOT / "data/research/talkmoves_recognizer_sweep.json")
    parser.add_argument(
        "--baseline", type=Path,
        default=ROOT / "data/research/talkmoves_recognizer_sweep.json",
        help="stored per-sentence sweep used by episode mode")
    parser.add_argument("--sample", type=Path, default=None,
                        help="write recognized fragments for a reading pass; "
                             "keep this OUTSIDE the repository (CC BY-NC-SA)")
    args = parser.parse_args()

    rows = student_sentences(args.transcripts)
    transcripts = {t for t, _, _ in rows}
    print(f"student sentences: {len(rows)} across {len(transcripts)} transcripts",
          flush=True)
    if not rows:
        print("FAIL: no student sentences read", file=sys.stderr)
        return 1

    if args.mode == "episode":
        if not args.baseline.exists():
            print(f"FAIL: sentence baseline {args.baseline} not found",
                  file=sys.stderr)
            return 1
        if (args.transcripts
                and args.output.resolve() == args.baseline.resolve()):
            print("FAIL: a limited episode run must use --output outside the "
                  "stored full-corpus baseline", file=sys.stderr)
            return 1
        baseline = json.loads(args.baseline.read_text(encoding="utf-8"))
        required = {
            "student_sentences", "transcripts", "candidate_support_levels"
        }
        missing = sorted(required - baseline.keys())
        if missing:
            print("FAIL: sentence baseline lacks "
                  f"{', '.join(missing)}", file=sys.stderr)
            return 1
        measured = episode_payload(
            rows, baseline, args.batch, args.timeout)
        payload = dict(baseline)
        payload["episode"] = measured
        args.output.parent.mkdir(parents=True, exist_ok=True)
        args.output.write_text(
            json.dumps(payload, indent=2, sort_keys=True) + "\n",
            encoding="utf-8")
        support = measured["candidate_support_levels"]
        print()
        print(f"student utterances       : {len(rows)}")
        print(f"transcript episodes      : {measured['episode_count']}")
        print(f"candidate support levels : {support}")
        print("distinct machines >=3    : "
              f"{measured['distinct_machines_reaching_three_or_more_ordered_actions']}")
        print("distinct machines accept : "
              f"{measured['distinct_machines_reaching_accepting_state']}")
        try:
            display_output = args.output.relative_to(ROOT)
        except ValueError:
            display_output = args.output
        print(f"wrote {display_output}")
        return 0

    fired = 0
    by_machine: collections.Counter = collections.Counter()
    by_family: collections.Counter = collections.Counter()
    by_support: collections.Counter = collections.Counter()
    by_machine_substantive: collections.Counter = collections.Counter()
    sentence_support: collections.Counter = collections.Counter()
    uncovered_math_language: collections.Counter = collections.Counter()
    substantive = 0
    per_transcript: collections.Counter = collections.Counter()
    sample: list[dict] = []

    for start in range(0, len(rows), args.batch):
        chunk = rows[start:start + args.batch]
        results = recognize([s for _, _, s in chunk], args.timeout)
        for (transcript, _speaker, sentence), line in zip(chunk, results):
            if line.strip() == "-":
                uncovered_math_language.update(divergence_matches(sentence))
                continue
            fired += 1
            per_transcript[transcript] += 1
            machines = []
            best_support = "none"
            for token in line.split():
                parts = token.split("|")
                if len(parts) != 5 or "/" not in parts[0]:
                    continue
                machine, support, matched, expected, confidence = parts
                by_machine[machine] += 1
                by_family[machine.split("/")[0]] += 1
                by_support[support] += 1
                if int(matched) >= 2:
                    substantive += 1
                    by_machine_substantive[machine] += 1
                if support == "partial_trace" or best_support == "partial_trace":
                    best_support = "partial_trace"
                elif best_support == "none":
                    best_support = support
                machines.append({"machine": machine, "support": support,
                                 "matched": int(matched),
                                 "expected": int(expected),
                                 "confidence": float(confidence)})
            sentence_support[best_support] += 1
            if args.sample and len(sample) < 400:
                sample.append({"transcript": transcript, "sentence": sentence,
                               "machines": machines})
        print(f"  {min(start + args.batch, len(rows))}/{len(rows)}"
              f"  fired={fired}", flush=True)

    payload = {
        "corpus": "TalkMoves (Sumner lab), CC BY-NC-SA 4.0; derived counts only",
        "student_sentences": len(rows),
        "transcripts": len(transcripts),
        "sentences_with_a_candidate": fired,
        "abstention_rate": round(1 - fired / len(rows), 4),
        "distinct_machines_fired": len(by_machine),
        "candidate_support_levels": dict(by_support.most_common()),
        "partial_trace_to_lexical_hint_ratio": round(
            by_support["partial_trace"] / by_support["lexical_hint"], 6
        ) if by_support["lexical_hint"] else None,
        "sentences_by_best_support": dict(sentence_support.most_common()),
        "uncovered_math_language": [
            {
                "pattern": pattern,
                "canonical_action": action,
                "abstained_sentences": count,
            }
            for (pattern, action), count
            in uncovered_math_language.most_common()
        ],
        "candidates_matching_two_or_more_actions": substantive,
        "distinct_machines_with_two_or_more_matched": len(by_machine_substantive),
        "top_machines_two_or_more_matched":
            dict(by_machine_substantive.most_common(25)),
        "by_family": dict(by_family.most_common()),
        "top_machines": dict(by_machine.most_common(40)),
        "transcripts_with_no_recognition":
            len(transcripts) - len(per_transcript),
    }
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(payload, indent=2, sort_keys=True) + "\n",
                           encoding="utf-8")
    if args.sample:
        args.sample.parent.mkdir(parents=True, exist_ok=True)
        args.sample.write_text(
            "\n".join(json.dumps(r, sort_keys=True) for r in sample) + "\n",
            encoding="utf-8")
        print(f"wrote {len(sample)} fragments to {args.sample} "
              "(licence: keep out of the repository)")

    print()
    print(f"student sentences        : {len(rows)}")
    print(f"recognized something     : {fired}")
    print(f"abstained                : {len(rows) - fired} "
          f"({payload['abstention_rate']:.1%})")
    print(f"distinct machines fired  : {len(by_machine)} of 232")
    print(f"candidate support levels : {dict(by_support.most_common())}")
    print(f"candidates matching >=2  : {substantive}")
    print(f"  over machines          : {len(by_machine_substantive)}")
    print(f"transcripts with nothing : {payload['transcripts_with_no_recognition']}"
          f" of {len(transcripts)}")
    try:
        display_output = args.output.relative_to(ROOT)
    except ValueError:
        display_output = args.output
    print(f"wrote {display_output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
